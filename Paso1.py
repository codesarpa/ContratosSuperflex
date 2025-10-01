import openpyxl
from openpyxl.styles import PatternFill
import time
import tkinter as tk
from tkinter import filedialog, messagebox
from selenium import webdriver
# from selenium.webdriver.firefox.service import Service
# from selenium.webdriver.firefox.options import Options
from Paso2 import IrPaginaContratos, Login
from Paso3 import ConsultarDocumento, GuardarContrato, GrupoPlanes
#import Paso2
#import Paso3
# Función para seleccionar archivo
def seleccionar_archivo():
    ruta_archivo = filedialog.askopenfilename(
        title="Selecciona el archivo de Excel",
        filetypes=[("Archivos de Excel", "*.xlsx")]
    )
    if ruta_archivo:
        ruta_archivo_entry.config(state=tk.NORMAL)
        ruta_archivo_entry.delete(0, tk.END)
        ruta_archivo_entry.insert(0, ruta_archivo)
        ruta_archivo_entry.config(state=tk.DISABLED)
    return ruta_archivo

# Función para limpiar la caja de texto y la ruta seleccionada
def limpiar_ruta():
    ruta_archivo_entry.config(state=tk.NORMAL)
    ruta_archivo_entry.delete(0, tk.END)
    ruta_archivo_entry.config(state=tk.DISABLED)

# Función para ejecutar el proceso
def ejecutar_proceso():
    ruta_archivo = ruta_archivo_entry.get()

    if not ruta_archivo:
        messagebox.showwarning("Advertencia", "Por favor, selecciona un archivo antes de ejecutar el proceso.")
        return

    try:
        libro = openpyxl.load_workbook(ruta_archivo)
        hoja = libro.active

        relleno_verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        driver = webdriver.Firefox()
        Login(driver)
        for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=hoja.max_column):
            for celda in fila:
                time.sleep(1)
                celda.fill = relleno_verde
                valor_celda = str(celda.value)
                print(f"RESULTADO {valor_celda}")
                ConsultarDocumento(valor_celda)
                GrupoPlanes()
                GuardarContrato(valor_celda)
                IrPaginaContratos()
                libro.save(ruta_archivo)

        #libro.save(ruta_archivo)
        messagebox.showinfo("Éxito", "El proceso se completó y el archivo se ha actualizado.")
        limpiar_ruta()
    except Exception as e:
        print("Error", f"Se produjo un error {valor_celda } al procesar el archivo: {str(e)}")
        #messagebox.showerror("Error", f"Se produjo un error al procesar el archivo: {str(e)}")

# Crear la ventana principal
root = tk.Tk()
root.title("Procesador de Excel")

# Etiqueta para seleccionar archivo
tk.Label(root, text="Seleccionar archivo:").grid(row=0, column=0, padx=10, pady=10)

# Caja de texto para mostrar la ruta del archivo
ruta_archivo_entry = tk.Entry(root, width=50, state=tk.DISABLED)
ruta_archivo_entry.grid(row=0, column=1, padx=10, pady=10)

# Botón para seleccionar archivo
boton_seleccionar = tk.Button(root, text="Seleccionar archivo", command=seleccionar_archivo)
boton_seleccionar.grid(row=0, column=2, padx=10, pady=10)

# Botón para limpiar la ruta seleccionada
boton_limpiar = tk.Button(root, text="Limpiar ruta", command=limpiar_ruta)
boton_limpiar.grid(row=1, column=2, padx=10, pady=10)

# Botón para ejecutar el proceso
boton_ejecutar = tk.Button(root, text="Ejecutar proceso", command=ejecutar_proceso)
boton_ejecutar.grid(row=2, column=0, columnspan=3, pady=20)

# Iniciar el loop de la aplicación
root.mainloop()
