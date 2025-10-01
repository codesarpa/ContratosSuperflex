import tkinter as tk
import openpyxl
from openpyxl.styles import PatternFill
from tkinter import filedialog, messagebox
from selenium import webdriver
import time
import telebot
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

def seleccionar_archivo():
    ruta_archivo = filedialog.askopenfilename(
        title="Selecciona el archivo de Excel",
        filetypes=[("Archivos de Excel", "*.xlsx")]
    )
    if ruta_archivo:
        ruta_archivo_entry.config(state=tk.NORMAL)
        ruta_archivo_entry.delete(0, tk.END)
        ruta_archivo_entry.insert(0, ruta_archivo)
        ruta_archivo_entry.config(state=tk.DISABLED)
    return ruta_archivo

def limpiar_ruta():
    ruta_archivo_entry.config(state=tk.NORMAL)
    ruta_archivo_entry.delete(0, tk.END)
    ruta_archivo_entry.config(state=tk.DISABLED)

def ejecutar_proceso():
    ruta_archivo = ruta_archivo_entry.get()

    if not ruta_archivo:
        messagebox.showwarning("Advertencia", "Por favor, selecciona un archivo antes de ejecutar el proceso.")
        return

    try:
        libro = openpyxl.load_workbook(ruta_archivo)
        hoja = libro.active

        relleno_verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        driver = webdriver.Firefox()
        Login(driver)
        for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=hoja.max_column):
            for celda in fila:
                time.sleep(1)
                celda.fill = relleno_verde
                valor_celda = str(celda.value)
                print(f"RESULTADO {valor_celda}")
                ConsultarDocumento(valor_celda, driver)
                GrupoPlanes(driver)
                GuardarContrato(valor_celda)
                IrPaginaContratos(driver)
                libro.save(ruta_archivo)

        #libro.save(ruta_archivo)
        messagebox.showinfo("Éxito", "El proceso se completó y el archivo se ha actualizado.")
        limpiar_ruta()
    except Exception as e:
        print("Error", f"Se produjo un error {valor_celda } al procesar el archivo: {str(e)}")
        #messagebox.showerror("Error", f"Se produjo un error al procesar el archivo: {str(e)}")

def IrPaginaContratos(driver):

    #driver.get("http://10.1.1.22:8181/BusinessNET-WEB/XHTML/azar/adminventa/contratopersonas.xhtml")
    driver.get("http://10.167.32.73:8130/BusinessNET-WEB/XHTML/azar/adminventa/contratopersonas.xhtml")
    print("Hola")
    # Esperar hasta que un elemento clave esté presente en la página
    elemento_clave = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "tbwContratoPersonas:frmPersonas:txtPersonaDocumentoPrincipal"))
    )
    print(elemento_clave)
    print("La página se ha cargado completamente.")
    time.sleep(3)

def Login(driver):
    try:        
        # Abrir la página de inicio de sesión
        #driver.get("http://10.1.1.22:8181/BusinessNET-WEB/XHTML/general/login.xhtml")
        # PRODUCCION driver.get("http://10.167.32.73:8130/BusinessNET-WEB/XHTML/general/login.xhtml")
        driver.get("http://10.1.1.22:8181/BusinessNET-WEB/XHTML/general/login.xhtml") #PRUEBAS
        time.sleep(3)
        driver.maximize_window()
        # Esperar hasta que el campo de usuario esté disponible
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "idFormLogin:user"))
        )

        # Ingresar el usuario
        username_field = driver.find_element(By.ID, "idFormLogin:user")
        username_field.send_keys("CP1121900795")

        # Ingresar la contraseña
        password_field = driver.find_element(By.ID, "idFormLogin:password")
        password_field.send_keys("795CP")
        print(password_field)
        # Hacer clic en el botón de ingresar
        login_button = driver.find_element(By.ID, "idFormLogin:ingresar")
        login_button.click()

        time.sleep(10)
        # Redireccionar a la página deseada después del inicio de sesión
        IrPaginaContratos(driver)
        time.sleep(5)
        # Aquí puedes continuar interactuando con la página después de la redirección

    except Exception as e:
        print("Error al iniciar sesión:", e)
        #driver.quit()
# Crear la ventana principal

def ConsultarDocumento(cedula, driver):
    # Esperar a que el campo esté presente
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located(
            (By.ID, "tbwContratoPersonas:frmPersonas:txtPersonaDocumentoPrincipal"))
    )

    # Encontrar el campo de texto y enviar el número de documento
    campo_documento = driver.find_element(
        By.ID, "tbwContratoPersonas:frmPersonas:txtPersonaDocumentoPrincipal")
    campo_documento.send_keys(cedula)
    campo_documento.send_keys(Keys.ENTER)

    # Esperar a que el campo de nombre tenga un valor distinto de vacío
    WebDriverWait(driver, 10).until(
        lambda d: d.find_element(
            By.ID, "tbwContratoPersonas:frmPersonas:txtPersonaNombre").get_attribute('value') != ""
    )
    # Hacer clic en el enlace "Personal ventas"
    enlace_personal_ventas = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.LINK_TEXT, "Personal ventas"))
    )
    enlace_personal_ventas.click()
    time.sleep(5)

def GrupoPlanes(driver):
    boton = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.ID, "tbwContratoPersonas:frmPersonasVenta:txtCodgrupoplanes"))
    )
    boton.click()
    time.sleep(2)

    input_box = driver.find_element(By.ID, "tbwContratoPersonas:frmPersonasVenta:txtCodgrupoplanes")

    input_box.clear()  # Limpia la caja de texto (opcional)
    input_box.send_keys("618")
    time.sleep(1)
    input_box.send_keys(Keys.RETURN)

    time.sleep(3)

def ValidarMensaje(driver):
    try:
        # Busca el elemento con la clase específica
        elemento = driver.find_element_by_css_selector(
            '.ui-growl-item-container.ui-state-highlight.ui-corner-all')

        # Verifica si el elemento es visible
        if elemento.is_displayed():
            print("El elemento está visible en pantalla.")
        else:
            print("El elemento existe pero no es visible en pantalla.")
    except NoSuchElementException:  # type: ignore
        print("El elemento no existe en el DOM.")

folder_path = './fotosPDV/'

def vaciarCarpeta():
    
    # Iterar sobre todos los archivos en la carpeta
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)  # Eliminar archivos o enlaces simbólicos
            elif os.path.isdir(file_path):
                os.rmdir(file_path)  # Eliminar directorios vacíos
        except Exception as e:
            print(f'Error al eliminar {file_path}. Razón: {e}')


def CapturaPantalla(cedula, driver):
    # Captura la pantalla completa
    screenshot_path = f"{folder_path}{cedula}.png"
    driver.save_screenshot(screenshot_path)


def EnvioFoto(cedula):

    bot = telebot.TeleBot("7241959128:AAEVyjfp1HPF1Ytpwe2gLNoSYQU3ZllgVx0")
    foto = open(f"{folder_path}{cedula}.png", "rb")
    bot.send_photo(7411433556, foto, f"<b>{cedula}!!</b>", parse_mode="html")


def GuardarContrato(cedula, driver):
    # click en el botón guardar
    boton_guardar = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "#tbwContratoPersonas\\:frmPersonasVenta\\:btnGuardarContratosventa")))
    # boton_guardar.click()
    action = ActionChains(driver)
    action.double_click(boton_guardar).perform()
    time.sleep(3)
    loading_element = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, "j_idt20"))
    )

    # Bucle para esperar mientras el elemento tiene la clase que indica que está visible y cargando
    while True:
        # Obtiene la clase del elemento
        element_class = loading_element.get_attribute("class")

        # Verifica si el elemento sigue visible y cargando
        if "ui-overlay-visible" in element_class:
            time.sleep(2)  # Espera 2 segundos antes de volver a verificar
        else:
            break

    time.sleep(10)
    print("salir ciclo")
    #EXTRAS
    CapturaPantalla(f"{cedula}")
    time.sleep(5)
    EnvioFoto(f"{cedula}")
    time.sleep(1)
    vaciarCarpeta()
    time.sleep(5)        

root = tk.Tk()
root.title("Procesador de Excel")

# Etiqueta para seleccionar archivo
tk.Label(root, text="Seleccionar archivo:").grid(row=0, column=0, padx=10, pady=10)

# Caja de texto para mostrar la ruta del archivo
ruta_archivo_entry = tk.Entry(root, width=50, state=tk.DISABLED)
ruta_archivo_entry.grid(row=0, column=1, padx=10, pady=10)

# Botón para seleccionar archivo
boton_seleccionar = tk.Button(root, text="Seleccionar archivo", command=seleccionar_archivo)
boton_seleccionar.grid(row=0, column=2, padx=10, pady=10)

# Botón para limpiar la ruta seleccionada
boton_limpiar = tk.Button(root, text="Limpiar ruta", command=limpiar_ruta)
boton_limpiar.grid(row=1, column=2, padx=10, pady=10)

# Botón para ejecutar el proceso
boton_ejecutar = tk.Button(root, text="Ejecutar proceso", command=ejecutar_proceso)
boton_ejecutar.grid(row=2, column=0, columnspan=3, pady=20)

# Iniciar el loop de la aplicación
root.mainloop()