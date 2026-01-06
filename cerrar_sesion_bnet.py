from httpcore import TimeoutException
import openpyxl
from config import *
from openpyxl.styles import PatternFill
from tkinter import filedialog, messagebox
from selenium import webdriver
import time
import telebot
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

folder_path = './fotosPDV/'

def CapturaPantalla(cedula):
    global driver
    # Captura la pantalla completa
    screenshot_path = f"{folder_path}{cedula}.png"
    logging.info(f"Captura Pantalla: {screenshot_path}")
    print(f"Captura Pantalla: {screenshot_path}")
    log_cedulas(f"Captura Pantalla: {screenshot_path}")
    driver.save_screenshot(screenshot_path)

def EnvioFoto(cedula, caption):
    #TELEGRAM_BOT_TOKEN = '7241959128:AAEVyjfp1HPF1Ytpwe2gLNoSYQU3ZllgVx0'
    #TELEGRAM_CHAT_ID = '-1002227166533'
    # GRUPO ROBOT CENSEL
    TELEGRAM_BOT_TOKEN = '8058959859:AAFIiLBcd4JI2DSrnEo2sKzFFCRikFctv2k'
    TELEGRAM_CHAT_ID = '7411433556'
    bot = telebot.TeleBot(TELEGRAM_BOT_TOKEN)
    foto = open(f"{folder_path}{cedula}.png", "rb")
    bot.send_photo(chat_id=TELEGRAM_CHAT_ID, photo=foto, caption=f"<b>{caption}!!</b>", parse_mode="html")
    logging.info("Envió imagen al Telegram")
    print("Envió imagen al Telegram")
    log_cedulas("Envió imagen al Telegram")

def IrPaginaContratos():
    global driver
    time.sleep(5)
    try:
        #driver.get("http://10.1.1.22:8181/BusinessNET-WEB/XHTML/azar/adminventa/contratopersonas.xhtml")
        driver.get("http://10.167.32.73:8130/BusinessNET-WEB/XHTML/azar/adminventa/contratopersonas.xhtml")
        # PRUEBAS driver.get("http://10.1.1.22:8181/BusinessNET-WEB/XHTML/azar/adminventa/contratopersonas.xhtml")
        logging.info("Ingreso a contratopersonas")
        log_cedulas("Ingreso a contratopersonas")
        
        # Esperar hasta que un elemento clave esté presente en la página
        elemento_clave = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "tbwContratoPersonas:frmPersonas:txtPersonaDocumentoPrincipal"))
        )

        logging.info(f"Se encuentra el campo dcto. principal")
        log_cedulas("Se encuentra el campo dcto. principal")
        print(f"Se encuentra el campo dcto. principal")
        logging.info("La pagina se carga correctamente")
        log_cedulas("La pagina se carga correctamente")
        print("La pagina se carga correctamente")
        
        time.sleep(3)
    except Exception as e:
        logging.error(f"Ocurrio un error al ingresar a contratos personas: {e}")
        print(f"Ocurrio un error al ingresar a contratos personas: {e}")
        log_cedulas(f"Ocurrio un error al ingresar a contratos personas: {e}")
        
def ConsultarDocumento(cedula):
    global driver
    try:
        logging.info(f"Se consulta el documento: {cedula}")
        log_cedulas(f"Se consulta el documento: {cedula}")
        print(f"Se consulta el documento: {cedula}")
        # Esperar a que el campo esté presente
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.ID, "tbwContratoPersonas:frmPersonas:txtPersonaDocumentoPrincipal"))
        )

        # Encontrar el campo de texto y enviar el número de documento
        campo_documento = driver.find_element(
            By.ID, "tbwContratoPersonas:frmPersonas:txtPersonaDocumentoPrincipal")
        campo_documento.send_keys(cedula)
        campo_documento.send_keys(Keys.ENTER)

        try:
        # Esperar a que el campo de nombre tenga un valor distinto de vacío
            WebDriverWait(driver, 10).until(
            lambda d: d.find_element(By.ID, "tbwContratoPersonas:frmPersonas:txtPersonaNombre").get_attribute('value') != "")
            campo_nombre = driver.find_element(By.ID, "tbwContratoPersonas:frmPersonas:txtPersonaNombre")
            valor_nombre = campo_nombre.get_attribute("value")
            logging.info(f"El campo nombre es:{valor_nombre}, de la cedula: {cedula}")
            log_cedulas(f"El campo nombre es:{valor_nombre}, de la cedula: {cedula}")
            print(f"El campo nombre es:{valor_nombre}, de la cedula: {cedula}")
        except Exception as e:  
            logging.error(f"No se pudo obtener el campo nombre: {str(e)}")
            log_cedulas(f"No se pudo obtener el campo nombre: {str(e)}")
            print(f"No se pudo obtener el campo nombre: {str(e)}")
            campo_nombre = None
            logging.error(f"El campo nombre es:{campo_nombre}")
            log_cedulas(f"El campo nombre es:{campo_nombre}")
            if campo_nombre == None:
                CapturaPantalla(f"{cedula}")
                EnvioFoto(f"{cedula}")
                logging.info("La cedula no es valida, por favor verificar")
                log_cedulas("La cedula no es valida, por favor verificar")
                print("La cedula no es valida, por favor verificar")
            return False
                
        # Hacer clic en el enlace "Personal ventas"
        enlace_personal_ventas = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.LINK_TEXT, "Personal ventas"))
        )
        enlace_personal_ventas.click()
        logging.info("Se encontro el enlace: Personal ventas")
        log_cedulas("Se encontro el enlace: Personal ventas")
        print("Se encontro el enlace: Personal ventas")
        time.sleep(5)
        return True
    except Exception as e:
        logging.error(f"Ocurrio un error al consultar el documento: {e}")
        print(f"Ocurrio un error al consultar el documento: {e}")

def GrupoPlanes():
    global driver
    try:
        boton = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.ID, "tbwContratoPersonas:frmPersonasVenta:txtCodgrupoplanes"))
        )
        boton.click()
        logging.info("Se encuentra el campo Grupo de Planes")
        print("Se encuentra el campo Grupo de Planes")
        log_cedulas("Se encuentra el campo Grupo de Planes")
        time.sleep(2)

        input_box = driver.find_element(By.ID, "tbwContratoPersonas:frmPersonasVenta:txtCodgrupoplanes")
        input_box.clear()  # Limpia la caja de texto (opcional)
        input_box.send_keys("620") #619 #200 #620 #111
        logging.info("Se escribe en el campo ")
        print("Se escribe en el campo ")
        log_cedulas("Se escribe en el campo ")
        
        time.sleep(1)
        input_box.send_keys(Keys.RETURN)

        time.sleep(3)
    except Exception as e:
        logging.error(f"Ocurrio un error en Grupo Planes {e}")
        print(f"Ocurrio un error en Grupo Planes {e}")
        log_cedulas(f"Ocurrio un error en Grupo Planes {e}")

def GuardarContrato(cedula):
    global driver
    try:
        # click en el botÃ³n guardar
        boton_guardar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#tbwContratoPersonas\\:frmPersonasVenta\\:btnGuardarContratosventa")))
        # boton_guardar.click()
        logging.info(f"Se encontro el boton guardar: {boton_guardar}")
        print(f"Se encontro el boton guardar: {boton_guardar}")
        log_cedulas(f"Se encontro el boton guardar: {boton_guardar}")
        driver.execute_script("arguments[0].scrollIntoView(true);", boton_guardar)
        time.sleep(2)  # Pequena pausa para que el scroll termine
        action = ActionChains(driver)
        action.double_click(boton_guardar).perform()
        time.sleep(3)
        
        loading_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "j_idt20"))
        )
        logging.info(f"Se encontro el elemento de carga: {loading_element}")
        print(f"Se encontro el elemento de carga: {loading_element}")
        logging.info("Cargando... esperando que termine el proceso.")
        print("Cargando... esperando que termine el proceso.")
        log_cedulas("Cargando... esperando que termine el proceso.")
        # Bucle para esperar mientras el elemento tiene la clase que indica que estÃ¡ visible y cargando
        while True:
            # Obtiene la clase del elemento
            element_class = loading_element.get_attribute("class")

            # Verifica si el elemento sigue visible y cargando
            if "ui-overlay-visible" in element_class:
                time.sleep(2)  # Espera 2 segundos antes de volver a verificar
            else:
                break
        
        time.sleep(5)
        print("fuera del while")
        #EXTRAS
        CapturaPantalla(f"{cedula}")
        time.sleep(5)
        EnvioFoto(f"{cedula}", f"{cedula}")
        time.sleep(1)
        # vaciarCarpeta()
        time.sleep(5)
        logging.info("Final proceso GuardarContrato")   
        print("Final proceso GuardarContrato")   
        
        
        # Esperar hasta 10 segundos por el mensaje de éxito
        try:
            mensaje_exito = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//p[text()='Registro actualizado exitosamente.']"))
            )
            mensaje_exito = mensaje_exito.get_attribute("value")
            # celda.fill = relleno_verde
            logging.info("Proceso exitoso: Registro actualizado correctamente.")
            print("Proceso exitoso: Registro actualizado correctamente.")
            logging.info(f"Mensaje_exito: {mensaje_exito}.")
            print(f"Mensaje_exito: {mensaje_exito}.")
            log_cedulas(f"Mensaje_exito: {mensaje_exito}.")
            return True
        except TimeoutException:
            logging.error("Error: No se encontro el mensaje de exito. Puede haber fallado el proceso.")
            print("Error: No se encontro el mensaje de exito. Puede haber fallado el proceso.")
            return False
        
    except Exception as e:
        logging.error(f"Ocurrio un error al guardar los cambios {e}")
        print(f"Ocurrio un error al guardar los cambios {e}")

def login():
    global driver
    try:
        # logging.info("SE INICIA EL PROCESO.")
        driver = webdriver.Firefox()        
        time.sleep(2)
        # Abrir la página de inicio de sesión
        #driver.get("http://10.1.1.22:8181/BusinessNET-WEB/XHTML/general/login.xhtml")
        driver.get("http://10.167.32.73:8130/BusinessNET-WEB/XHTML/general/login.xhtml")
        # driver.get("http://10.1.1.25:8181/BusinessNET-WEB/XHTML/general/login.xhtml") #PRUEBAS
        logging.info("Se abre Bnet")
        print("Se abre Bnet")
        time.sleep(3)
        driver.maximize_window()
        # Esperar hasta que el campo de usuario esté disponible
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "idFormLogin:user"))
        )
        logging.info("Se Encuentra el campo usuario")
        print("Se Encuentra el campo usuario")
        # Ingresar el usuario
        username_field = driver.find_element(By.ID, "idFormLogin:user")
        username_field.send_keys("CP1029980182")
        time.sleep(2)
        # Ingresar la contraseña
        password_field = driver.find_element(By.ID, "idFormLogin:password")
        # PRUEBAS password_field.send_keys("795CP")
        # password_field.send_keys("ve56gv79")
        password_field.send_keys("Stiff356")
        logging.info("Se encuentra el campo contraseña")
        print("Se encuentra el campo contraseña")
        # Hacer clic en el botón de ingresar
        login_button = driver.find_element(By.ID, "idFormLogin:ingresar")
        login_button.click()
        logging.info("Se encuentra el boton Ingresar")
        print("Se encuentra el boton Ingresar")

        time.sleep(10)
        
        # #Redirigir a la ruta contratopersonas
        # driver = webdriver.Firefox()
        # driver.get("http://10.1.1.22:8181/BusinessNET-WEB/XHTML/azar/adminventa/contratopersonas.xhtml")
        # print("Hola")
        # # Esperar hasta que un elemento clave esté presente en la página
        # elemento_clave = WebDriverWait(driver, 10).until(
        #     EC.presence_of_element_located((By.ID, "tbwContratoPersonas:frmPersonas:txtPersonaDocumentoPrincipal"))
        # )
        
        # print(elemento_clave)
        # print("La página se ha cargado completamente.")
        # # Redireccionar a la página deseada después del inicio de sesión
        # # IrPaginaContratos(driver)
        # time.sleep(5)
        # # Aquí puedes continuar interactuando con la página después de la redirección
        return driver
    except Exception as e:
        print("Error al iniciar sesión:", e)
        logging.error(f"Error al inciar sesion: {e}")
        #driver.quit()

def ingresar_menu_resetear_usuarios():
    global driver
    try:
        driver.get("http://10.167.32.73:8130/BusinessNET-WEB/XHTML/azar/logisticaventas/administracionusuarios.xhtml")
        logging.info("Ingreso a la administracion de usuarios para el reseteo del usuario")
        log_cedulas("Ingreso a la administracion de usuarios para el reseteo del usuario")

        time.sleep(5)
        #boton para cargar los usuarios, dar click antes de consultar cualquier usuario
        boton_ejecutar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.ID, "tbwAdmusuarios:frmusuarios:btnIncrementarUsuarios"))
            )
        boton_ejecutar.click()

        loading_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "j_idt20"))
            )

        while True:
            # Obtiene la clase del elemento
            element_class = loading_element.get_attribute("class")

            # Verifica si el elemento sigue visible y cargando
            if "ui-overlay-visible" in element_class:
                print("Cargando..")
                time.sleep(2)  # Espera 2 segundos antes de volver a verificar
            else:
                break

        logging.info("Se ingresa correctamente al modulo de resetear usuarios.")
        print("Se ingresa correctamente al modulo de resetear usuarios.")
        log_cedulas(f"Se ingresa correctamente al modulo de resetear usuarios.")
        return True
    except Exception as e:
        logging.error(f"Ocurrio un error al ingresar a la administracion de usuarios: {e}")
        print(f"Ocurrio un error al ingresar a la administracion de usuarios: {e}")
        log_cedulas(f"Ocurrio un error al ingresar a la administracion de usuarios: {e}")
        return False
    
def consultar_usuario_cambiar_estado(cedula):
    global driver
    logging.info(f"Se consulta la siguiente cedula: {cedula}")
    #boton de consultar usuario para resetear
    
    time.sleep(5)
    input_consultar_usuario = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.ID, "tbwAdmusuarios:frmusuarios:dtbusuarios:loginusr:filter"))
        )
    input_consultar_usuario.click()
    #consultar la cedula
    input_consultar_usuario.send_keys(cedula)        
    input_consultar_usuario.send_keys(Keys.ENTER)        

    time.sleep(5)
    #boton para editar cedula consultada (logo de un lapiz)
    btn_editar_cedula_consultada = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.ID, "tbwAdmusuarios:frmusuarios:dtbusuarios:0:btneditarusuarios"))
        )
    btn_editar_cedula_consultada.click()

    time.sleep(5)
    #boton de reseteo estado
    btn_reseteo_estado = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.ID, "tbwAdmusuarios:frmdlgAdminusuarios:j_idt171"))
        )
    btn_reseteo_estado.click()

    time.sleep(3)
    mensaje_respuesta = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "/html/body/div[12]/div/div/div[2]/p"))
    )
    print(mensaje_respuesta.text)
    logging.info(f"se da click en el boton de reseto estado y se valida el siguiente mensaje: {mensaje_respuesta.text}")

    CapturaPantalla(f"{cedula}")
    EnvioFoto(f"{cedula}", f"reseteo de estado: {cedula} \n{mensaje_respuesta.text}")
    #dar click en boton guardar
    btn_guardar = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.ID, "tbwAdmusuarios:frmdlgAdminusuarios:btnguardarusuarios"))
        )
    btn_guardar.click()

    time.sleep(3)
    mensaje_respuesta_guardar = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, "/html/body/div[12]/div/div/div[2]/p"))
    )
    print(mensaje_respuesta_guardar.text)
    logging.info(f"se da click en el boton de guardar y se valida el siguiente mensaje: {mensaje_respuesta_guardar.text}")

    CapturaPantalla(f"{cedula}")
    EnvioFoto(f"{cedula}", f"guardar reseteo: {cedula} \n{mensaje_respuesta_guardar.text}")
    #primero click en la ruedita
    # driver.get("http://10.167.32.73:8130/BusinessNET-WEB/XHTML/azar/logisticaventas/administracionusuarios.xhtml")

logging.info("")
logging.info(F"INICIO EJECUCION DE CERRAR SESION {hora_inicio}")
# nombre_archivo = 'cedulas.xlsx'
nombre_archivo = 'cedulas.xlsx'
ruta_archivo = f'C:\ContratosSuperflex\{nombre_archivo}'
logging.info(F"Nombre del archivo seleccionado: {nombre_archivo}")
libro = openpyxl.load_workbook(ruta_archivo)
hoja = libro.active
relleno_verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
relleno_rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

driver = login()
if driver:
    for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=1, max_col=1):
        logging.info("Se inicia iteracion")
        print("Se inicia iteracion")
        for celda in fila:
            logging.info("Procesando celda")
            print("Procesando celda")
            if celda.value:
                hay_cedulas = True
                print(celda.value)
                valor_celda = str(celda.value)
                logging.info(f"Se inicia el proceso con la Cedula: {valor_celda}")
                log_cedulas("Se abre Bnet")
                time.sleep(1)
                consultar_usuario = ingresar_menu_resetear_usuarios()
                if consultar_usuario:
                    consultar_usuario_cambiar_estado(valor_celda)
                    IrPaginaContratos()
                    consulta_documento = ConsultarDocumento(valor_celda)
                    if consulta_documento == False:
                        logging.warning(f"Consulta fallida para cédula {valor_celda}. Continuando con la siguiente...")
                        print(f"Consulta fallida para cédula {valor_celda}. Continuando con la siguiente...")
                        continue
                    GrupoPlanes()
                    guardar_contrato = GuardarContrato(valor_celda)
                    if guardar_contrato == True:
                        celda.fill = relleno_verde
                        logging.info(f"Se valida el proceso de la cedula: {valor_celda} EXITOSO")
                    else:
                        celda.fill = relleno_rojo
                        logging.info(f"Se valida el proceso de la cedula: {valor_celda} FALLIDO")
                        cedulas_fallidas = valor_celda

                    libro.save(ruta_archivo)
                    logging.info("Se guarda el archivo excel correctamente")
                else:
                    print("Fallo al resetear el estado, se continua con la siguiente cedula")
                    logging.info("Fallo al resetear el estado, se continua con la siguiente cedula")
                    continue

    print("Finalizó el bucle de iteración sobre las cédulas.")
    print(f"hay_cedulas: {hay_cedulas}")            
    if not hay_cedulas:
        logging.info("Ya no se encuentran cédulas en el archivo Excel")
        logging.shutdown()
        print("antes del exit")
        exit()
